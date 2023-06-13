#Modules
import openpyxl, os
from time import sleep
from selenium import webdriver                                                                                          
from selenium.webdriver.common.keys import Keys                                                                         
from selenium.webdriver import ActionChains                                                                                                         
                 	
#Inputs
directory = r'REMOVED'
file = 'REMOVED.xlsx'                                                                                         	
sheet ='10'                  

#Input Conditions
#In the excel I use ... REMOVED
#This all lines up with a Macro I have used, I should thereotically fix this so it works on a non-editted Excel
                                                       
#Variables
os.chdir(directory)
workbook = openpyxl.load_workbook(file)[sheet]
rows = workbook.max_row
driver = webdriver.Edge(r'REMOVED')
                                                                                       
#Actions
print(f'ETAS to Process = {rows-1}')                                                                                                                                           
driver.get('REMOVED')
sleep(10)                                                                                                          

#Loop
for i in range(2,rows+1):
    driver.implicitly_wait(10) 
    job_number = workbook.cell(row=i, column=1).value
    eta = workbook.cell(row=i, column=4).value
    resource = workbook.cell(row=i, column=3).value
        
    #Job Number & Job Notes LookUp
    try:
        sleep(5)                                                                                                  
        driver.find_element_by_id("REMOVED").send_keys(job_number, Keys.ENTER)                          #job search                                                                    
        sleep(3)
        #REMOVED = driver.find_element_by_xpath('REMOVED')
        #REMOVED = do_not_sms.find_element_by_tag_name('REMOVED')
        #if not REMOVED.is_selected():
            #REMOVED.click()
        sleep(0.5)
        REMOVED = driver.find_element_by_id('REMOVED')
        sleep(1)
        for row in REMOVED.find_elements_by_css_selector('REMOVED'):
            if resource in row.text:
                raise_event = row
                break
        sleep(2)
        ActionChains(driver).context_click(raise_event).perform()
        sleep(1)
        driver.find_element_by_xpath('REMOVED').click()
        sleep(2)
        driver.find_element_by_id("REMOVED").send_keys(eta)                                   #add note, send excel value dependent on priority                                               
        sleep(5)                                                                                           
        driver.find_element_by_xpath('//*[@id="REMOVED"]').click()                              #save ETA                                                                                      
        print(f'{i-1}. {job_number} - SUCCESS')                                                                                                     
    except Exception as e:
        print(f'{i-1}. {job_number} - ERROR. Full Error Below')
        print(e)

